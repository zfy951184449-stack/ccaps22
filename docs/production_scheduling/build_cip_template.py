"""
生成「CIP 拓扑导入模板.xlsx」。改这里、跑 `python3 build_cip_template.py` 重新生成同目录的 xlsx。

设备「类型」取自系统「设备类型」字典(ps_equipment_type,可在系统里增删改)。本模板的类型下拉
只是「当前默认值」的快照、且为非阻断(可直接改写);导入时以系统活字典校验,不在字典/已停用会被拒。
若字典有大改,可把下面 TYPES 同步一下再重生成,但不是必须。
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

HEAD_FILL = PatternFill('solid', fgColor='1F4E79')
HEAD_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
REQ_FONT = Font(name='Arial', bold=True, color='FFFF00', size=11)
THIN = Side(style='thin', color='D9D9D9')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ALIGN = Alignment(horizontal='left', vertical='center')

wb = Workbook()

def add_name(name, ref):
    dn = DefinedName(name, attr_text=ref)
    try:
        wb.defined_names.add(dn)
    except AttributeError:
        wb.defined_names[name] = dn

def style_header(ws, headers, required):
    for c, (h, req) in enumerate(zip(headers, required), start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEAD_FILL
        cell.font = REQ_FONT if req else HEAD_FONT
        cell.alignment = ALIGN
        cell.border = BORDER
    ws.freeze_panes = 'A2'

def fill_rows(ws, rows):
    for r, row in enumerate(rows, start=2):
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = Font(name='Arial', size=11)
            cell.alignment = ALIGN
            cell.border = BORDER

# ── 说明 ──
ws_doc = wb.active
ws_doc.title = '说明'
doc_lines = [
    ('CIP 拓扑导入模板 · 填写说明', True),
    ('', False),
    ('1) 一个工作簿 = 一个设施。设施在导入页选择,本表不用填 facility。', False),
    ('2) 被 CIP 洗的对象有两类、平级:设备(罐/单元)和 管线(设备-设备的连接,如 pouA-PT),各自归属一个 CIP 站。', False),
    ('3) 房间/设备归属「组织」(组织编码=team 的编码(只到 team 层级),与排班同一套组织);设备组织/房间留空则随所在房间或上级设备。', False),
    ('3b) 设备可成树:pou 等「使用点/端口」填「上级设备编码」挂到母设备(罐/skid/配液系统)下,成为它的子设备;留空=顶层设备。上级可以是本表里后面才录的设备。', False),
    ('3c) 「类型」取自系统「设备类型」字典(可在系统里随时增删改);本下拉为当前默认值、可直接改写,导入时以系统活字典校验(不在字典或已停用会被拒)。', False),
    ('4) 引用一律用「编码」:设备的 上级设备编码/房间编码←房间表、CIP站编码←站表;管线的 起点/终点设备编码←设备表、CIP站编码←站表。', False),
    ('5) 设备「清洗方式」:CIP(在线清洗,要填 CIP站)/ 一次性(免洗,如一次性反应器)/ COP(离线)/ 其他。', False),
    ('     只有 CIP 设备才进 CIP 排程、才算尖峰;非 CIP 设备不填 CIP站。', False),
    ('5b) 清洗时序 = 3 个动作时长(分钟)× 4 个保持窗(小时)。动作:CIP(全清洗)/ RIP(淋洗在位,层析/UFDF 常以 RIP+SIP 替代罐式 CIP)/ SIP(灭菌)。', False),
    ('     保持窗:DHT(脏停放=变脏后须几小时内开洗)/ RHT(淋洗有效期=RIP 完须几小时内 SIP/被用)/ CHT(洁净有效期=CIP 完须几小时内被用)/ SHT(无菌有效期=SIP 完须几小时内被用)。', False),
    ('     只对要清洗的设备/管线填,不做的留空;管线也可 RIP+SIP(在线灭菌转移线)。', False),
    ('6) 带 * 的列必填(表头标黄)。枚举类下拉(清洗方式/洁净级别/类别/起算)填错会被拒;「类型」下拉为建议、可改写。', False),
    ('7) 填写顺序:先「站」「房间」→ 再「设备」→ 再「管线」。下游下拉来自上游已录的编码,自动增长。', False),
    ('8) 导入按「设施+编码」upsert:已存在则更新,不存在则新增;重复导入不会建重复。', False),
]
for r, (txt, head) in enumerate(doc_lines, start=1):
    cell = ws_doc.cell(row=r, column=1, value=txt)
    cell.font = Font(name='Arial', bold=head, size=14 if head else 11)
ws_doc.column_dimensions['A'].width = 112

# ── 选项(隐藏) ──
ws_opt = wb.create_sheet('选项')
ws_opt['A1'] = '设备类型'; ws_opt['B1'] = '效期类别'; ws_opt['C1'] = '起算基准'; ws_opt['D1'] = '清洗方式'; ws_opt['E1'] = '洁净级别'; ws_opt['F1'] = '组织编码'
# 与 database/migrations/20260628_equipment_type_dict.sql 的默认字典一致(快照,非阻断)
TYPES = ['生物反应器', '种子反应器', '配液罐', '缓冲液罐', '培养基罐', '储液罐', '中间罐', '移动罐',
         '离心机', '深层过滤器', '除菌过滤器', '层析 skid', '超滤 skid', '病毒灭活', '病毒过滤',
         '冻干机', '灌装机', '称量配制', '取样', 'CIP skid', 'SIP skid', '转移/管路', '其他']
CATS = ['培养基', '缓冲液', '清洗剂', '中间产物', '试剂', '设备洁净']
BASIS = ['产出后', '配制后', '清洗后']
CLEAN = ['CIP', '一次性', 'COP', '其他']
GRADE = ['A', 'B', 'C', 'D', 'CNC']
# 组织编码:只到 team 层级(organization_units 里 unit_type=TEAM 的 unit_code)。如有变动重新生成模板。
ORG = ['USP', 'DSP', 'SPI', 'MEDIA']
for i, v in enumerate(TYPES, start=2): ws_opt.cell(row=i, column=1, value=v)
for i, v in enumerate(CATS, start=2): ws_opt.cell(row=i, column=2, value=v)
for i, v in enumerate(BASIS, start=2): ws_opt.cell(row=i, column=3, value=v)
for i, v in enumerate(CLEAN, start=2): ws_opt.cell(row=i, column=4, value=v)
for i, v in enumerate(GRADE, start=2): ws_opt.cell(row=i, column=5, value=v)
for i, v in enumerate(ORG, start=2): ws_opt.cell(row=i, column=6, value=v)
ws_opt.sheet_state = 'hidden'

# ── 站 ──
ws_st = wb.create_sheet('站')
style_header(ws_st, ['站编码*', '站名称*', '组织编码', '容量', '备注'], [1,1,0,0,0])
fill_rows(ws_st, [['CIP-1','上游CIP站','USP',1,'组织编码=team编码'], ['CIP-2','下游CIP站','DSP',1,'']])
for col, w in zip('ABCDE', [16,22,12,8,28]): ws_st.column_dimensions[col].width = w

# ── 房间 ──
ws_rm = wb.create_sheet('房间')
style_header(ws_rm, ['房间编码*', '房间名称*', '组织编码', '洁净级别', '备注'], [1,1,0,0,0])
fill_rows(ws_rm, [
    ['R-1501','上游生产间A','USP','C','组织编码=team编码(只到team层级)'],
    ['R-1601','下游纯化间','DSP','C',''],
])
for col, w in zip('ABCDE', [16,22,14,10,26]): ws_rm.column_dimensions[col].width = w

# ── 设备 ──
# 时序列 = 3 动作时长(CIP/RIP/SIP·分钟,I/J/K)× 4 保持窗(DHT/RHT/CHT/SHT·小时,L/M/N/O)
ws_eq = wb.create_sheet('设备')
style_header(ws_eq, ['设备编码*', '设备名称*', '类型*', '清洗方式', '上级设备编码', '房间编码', '组织编码', 'CIP站编码', 'CIP时长(分钟)', 'RIP时长(分钟)', 'SIP时长(分钟)', 'DHT(小时)', 'RHT(小时)', 'CHT(小时)', 'SHT(小时)', '备注'], [1,1,1,0,0,0,0,0,0,0,0,0,0,0,0,0])
fill_rows(ws_eq, [
    ['PT','配液罐PT','配液罐','CIP','','R-1501','','CIP-1',90,'','',24,'',72,'','组织留空=随房间'],
    ['BIO-2000','2000L不锈钢反应器','生物反应器','CIP','','R-1501','','CIP-1',120,'',60,24,'',72,48,'CIP后SIP→填SHT'],
    ['SUB-50','一次性反应器','生物反应器','一次性','','R-1501','','','','','','','','','','一次性免CIP,站/时长全留空'],
    ['BUF-SYS','缓冲配液系统','配液罐','CIP','','R-1501','','CIP-1',60,'','',12,'',48,'','母系统:下面挂多个使用点'],
    ['pouA','使用点A','转移/管路','其他','BUF-SYS','','','','','','','','','','','子设备:房间/组织留空→随上级BUF-SYS'],
    ['AKTA-1','层析skid 1','层析 skid','CIP','','R-1601','','CIP-2','',40,30,24,8,'',12,'层析不做罐式CIP:RIP+SIP'],
])
for col, w in zip('ABCDEFGHIJKLMNOP', [16,22,12,12,14,14,12,12,13,13,13,11,11,11,11,24]): ws_eq.column_dimensions[col].width = w

# ── 管线 ──
# 时序列 = CIP/RIP/SIP·分钟(F/G/H)× DHT/RHT/CHT/SHT·小时(I/J/K/L);转移线可 RIP+SIP
ws_pl = wb.create_sheet('管线')
style_header(ws_pl, ['管线编码*', '管线名称*', '起点设备编码*', '终点设备编码*', 'CIP站编码*', 'CIP时长(分钟)', 'RIP时长(分钟)', 'SIP时长(分钟)', 'DHT(小时)', 'RHT(小时)', 'CHT(小时)', 'SHT(小时)', '备注'], [1,1,1,1,1,0,0,0,0,0,0,0,0])
fill_rows(ws_pl, [['pouA-PT','pouA到PT配液线','pouA','PT','CIP-1',45,'','',24,'',48,'','示例行']])
for col, w in zip('ABCDEFGHIJKLM', [16,22,16,16,14,13,13,13,11,11,11,11,22]): ws_pl.column_dimensions[col].width = w

# ── 物料效期 ──
ws_sl = wb.create_sheet('物料效期')
style_header(ws_sl, ['物料*', '类别*', '效期(小时)*', '起算基准', '备注'], [1,1,1,0,0])
fill_rows(ws_sl, [
    ['培养基','培养基',24,'配制后','示例行'],
    ['1M NaOH','清洗剂',168,'配制后',''],
    ['洁净反应器','设备洁净',72,'清洗后','这就是 CHT(洁净有效期)'],
])
for col, w in zip('ABCDE', [18,14,14,12,28]): ws_sl.column_dimensions[col].width = w

# ── 动态命名区域 ──
add_name('站编码', "OFFSET('站'!$A$2,0,0,MAX(1,COUNTA('站'!$A$2:$A$1000)),1)")
add_name('房间编码', "OFFSET('房间'!$A$2,0,0,MAX(1,COUNTA('房间'!$A$2:$A$1000)),1)")
add_name('设备编码', "OFFSET('设备'!$A$2,0,0,MAX(1,COUNTA('设备'!$A$2:$A$1000)),1)")
add_name('设备类型', "'选项'!$A$2:$A$24")
add_name('效期类别', "'选项'!$B$2:$B$7")
add_name('起算基准', "'选项'!$C$2:$C$4")
add_name('清洗方式', "'选项'!$D$2:$D$5")
add_name('洁净级别', "'选项'!$E$2:$E$6")
add_name('组织编码', "'选项'!$F$2:$F$5")

def dv(ws, formula, cells, stop=True, prompt=None):
    d = DataValidation(type='list', formula1=formula, allow_blank=True,
                       showErrorMessage=True, showInputMessage=bool(prompt))
    d.errorStyle = 'stop' if stop else 'warning'
    d.error = '请从下拉中选择有效值' if stop else '该值不在当前下拉,确认无误再继续(导入时以系统校验)'
    d.errorTitle = '无效值'
    if prompt:
        d.prompt = prompt; d.promptTitle = '提示'
    ws.add_data_validation(d); d.add(cells)

# 站:组织编码
dv(ws_st, '=组织编码', 'C2:C500', stop=False, prompt='team 编码(只到 team 层级)')
# 房间:组织编码 + 洁净级别
dv(ws_rm, '=组织编码', 'C2:C500', stop=False, prompt='team 编码(只到 team 层级)')
dv(ws_rm, '=洁净级别', 'D2:D500', stop=True)
# 设备:类型(取自系统字典,非阻断可改写)+ 清洗方式(枚举)+ 上级设备(←设备,自引用)+ 房间编码(←房间)+ 组织编码 + CIP站编码(←站)
dv(ws_eq, '=设备类型', 'C2:C500', stop=False, prompt='默认类型;不够用就在系统「设备类型」里加,这里可直接改写')
dv(ws_eq, '=清洗方式', 'D2:D500', stop=True)
dv(ws_eq, '=设备编码', 'E2:E500', stop=False, prompt='可选:pou 等使用点挂到母设备/skid;选「设备」里已录的编码,留空=顶层')
dv(ws_eq, '=房间编码', 'F2:F500', stop=False, prompt='选「房间」表里已录的房间编码;留空随上级设备')
dv(ws_eq, '=组织编码', 'G2:G500', stop=False, prompt='留空=随房间/上级;否则填 team 编码')
dv(ws_eq, '=站编码', 'H2:H500', stop=False, prompt='仅清洗方式=CIP 时填,选「站」表里已录的站编码')
# 管线:起点/终点(←设备)+ CIP站编码(←站)
dv(ws_pl, '=设备编码', 'C2:C500', stop=False, prompt='选「设备」表里已录的设备编码')
dv(ws_pl, '=设备编码', 'D2:D500', stop=False, prompt='选「设备」表里已录的设备编码')
dv(ws_pl, '=站编码', 'E2:E500', stop=False, prompt='选「站」表里已录的站编码')
# 物料效期:枚举
dv(ws_sl, '=效期类别', 'B2:B500', stop=True)
dv(ws_sl, '=起算基准', 'D2:D500', stop=True)
# 数字
cap = DataValidation(type='whole', operator='greaterThanOrEqual', formula1='1', allow_blank=True)
ws_st.add_data_validation(cap); cap.add('D2:D500')
hrs = DataValidation(type='whole', operator='greaterThanOrEqual', formula1='0', allow_blank=True)
ws_sl.add_data_validation(hrs); hrs.add('C2:C500')
# 清洗时序(非负整数):设备 CIP/RIP/SIP时长 + DHT/RHT/CHT/SHT = I..O;管线 = F..L
eqnum = DataValidation(type='whole', operator='greaterThanOrEqual', formula1='0', allow_blank=True)
ws_eq.add_data_validation(eqnum); eqnum.add('I2:O500')
plnum = DataValidation(type='whole', operator='greaterThanOrEqual', formula1='0', allow_blank=True)
ws_pl.add_data_validation(plnum); plnum.add('F2:L500')

order = ['说明', '站', '房间', '设备', '管线', '物料效期', '选项']
wb._sheets.sort(key=lambda s: order.index(s.title))

import os
out = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'CIP拓扑导入模板.xlsx')
wb.save(out)
print('saved', out)
